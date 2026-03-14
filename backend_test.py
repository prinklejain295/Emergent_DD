import requests
import sys
import json
import io
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook

class DueDateAPITester:
    def __init__(self, base_url="https://client-tax-notify.preview.emergentagent.com/api"):
        self.base_url = base_url
        self.token = None
        self.user_id = None
        self.organization_id = None
        self.client_id = None
        self.custom_service_id = None
        self.due_date_id = None
        self.tests_run = 0
        self.tests_passed = 0
        self.test_email = f"testorg{datetime.now().strftime('%H%M%S')}@duedate.com"

    def log_test(self, name, success, details=""):
        """Log test results"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
            print(f"✅ {name} - PASSED")
        else:
            print(f"❌ {name} - FAILED: {details}")
        return success

    def make_request(self, method, endpoint, data=None, files=None, expected_status=200):
        """Make HTTP request with proper headers"""
        url = f"{self.base_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'
        
        if files:
            headers.pop('Content-Type', None)  # Let requests set it for multipart

        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, files=files, headers=headers)
                else:
                    response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)
            elif method == 'PATCH':
                response = requests.patch(url, json=data, headers=headers)

            success = response.status_code == expected_status
            return success, response.json() if response.content else {}, response.status_code

        except Exception as e:
            return False, {"error": str(e)}, 0

    def test_1_register_with_organization(self):
        """Test 1.1: Register with Organization"""
        data = {
            "email": self.test_email,
            "password": "Test123!",
            "name": "Test Manager",
            "organization_name": "Test Accounting Firm"
        }
        
        success, response, status = self.make_request('POST', 'auth/register', data)
        
        if success and 'access_token' in response:
            self.token = response['access_token']
            self.user_id = response['user']['id']
            self.organization_id = response['user']['organization_id']
            return self.log_test("Register with Organization", True)
        else:
            return self.log_test("Register with Organization", False, f"Status: {status}, Response: {response}")

    def test_2_login(self):
        """Test 1.2: Login"""
        data = {
            "email": self.test_email,
            "password": "Test123!"
        }
        
        success, response, status = self.make_request('POST', 'auth/login', data)
        
        if success and 'access_token' in response:
            self.token = response['access_token']
            return self.log_test("Login", True)
        else:
            return self.log_test("Login", False, f"Status: {status}, Response: {response}")

    def test_3_get_service_types(self):
        """Test 2.1: Get All Service Types"""
        success, response, status = self.make_request('GET', 'service-types')
        
        if success and isinstance(response, dict):
            # Check required categories
            required_categories = ['federal', 'state', 'payroll', 'other', 'custom']
            has_all_categories = all(cat in response for cat in required_categories)
            
            # Check federal types count (should be 15+)
            federal_count = len(response.get('federal', []))
            has_enough_federal = federal_count >= 15
            
            # Check for specific required forms
            federal_types = response.get('federal', [])
            required_forms = ['Form 941', 'W-2', '1099-NEC']
            has_required_forms = any(any(form in ft for form in required_forms) for ft in federal_types)
            
            if has_all_categories and has_enough_federal and has_required_forms:
                return self.log_test("Get All Service Types", True, f"Found {federal_count} federal types")
            else:
                return self.log_test("Get All Service Types", False, 
                    f"Categories: {has_all_categories}, Federal count: {federal_count}, Required forms: {has_required_forms}")
        else:
            return self.log_test("Get All Service Types", False, f"Status: {status}, Response: {response}")

    def test_4_create_custom_service_type(self):
        """Test 2.2: Create Custom Service Type"""
        data = {
            "name": "Business License Renewal - California",
            "category": "custom"
        }
        
        success, response, status = self.make_request('POST', 'service-types', data)
        
        if success and 'id' in response:
            self.custom_service_id = response['id']
            return self.log_test("Create Custom Service Type", True)
        else:
            return self.log_test("Create Custom Service Type", False, f"Status: {status}, Response: {response}")

    def test_5_get_service_types_after_custom(self):
        """Test 2.3: Get Service Types After Adding Custom"""
        success, response, status = self.make_request('GET', 'service-types')
        
        if success and isinstance(response, dict):
            custom_types = response.get('custom', [])
            has_custom = "Business License Renewal - California" in custom_types
            
            if has_custom:
                return self.log_test("Get Service Types After Adding Custom", True)
            else:
                return self.log_test("Get Service Types After Adding Custom", False, 
                    f"Custom types: {custom_types}")
        else:
            return self.log_test("Get Service Types After Adding Custom", False, f"Status: {status}")

    def test_6_create_individual_client(self):
        """Test 3.1: Create Individual Client"""
        data = {
            "name": "Tech Startup Inc",
            "email": "contact@techstartup.com",
            "phone": "555-0123",
            "company": "Tech Startup Inc",
            "notes": "New tech company client"
        }
        
        success, response, status = self.make_request('POST', 'clients', data)
        
        if success and 'id' in response:
            self.client_id = response['id']
            return self.log_test("Create Individual Client", True)
        else:
            return self.log_test("Create Individual Client", False, f"Status: {status}, Response: {response}")

    def test_7_excel_bulk_upload(self):
        """Test 3.2: Excel Bulk Upload"""
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Email'
        ws['C1'] = 'Phone'
        ws['D1'] = 'Company'
        ws['E1'] = 'Notes'
        
        # Test data
        test_clients = [
            ['Client A', 'clienta@test.com', '555-1111', 'Company A', 'Notes A'],
            ['Client B', 'clientb@test.com', '555-2222', 'Company B', 'Notes B'],
            ['Client C', 'clientc@test.com', '555-3333', 'Company C', 'Notes C']
        ]
        
        for i, client in enumerate(test_clients, start=2):
            for j, value in enumerate(client, start=1):
                ws.cell(row=i, column=j, value=value)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {'file': ('test_clients.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response, status = self.make_request('POST', 'clients/upload-excel', files=files)
        
        if success and response.get('clients_created') == 3:
            return self.log_test("Excel Bulk Upload", True, f"Created {response['clients_created']} clients")
        else:
            return self.log_test("Excel Bulk Upload", False, f"Status: {status}, Response: {response}")

    def test_8_verify_bulk_import(self):
        """Test 3.3: Verify Bulk Import"""
        success, response, status = self.make_request('GET', 'clients')
        
        if success and isinstance(response, list):
            client_count = len(response)
            # Should have 4 clients (1 manual + 3 bulk)
            if client_count >= 4:
                return self.log_test("Verify Bulk Import", True, f"Found {client_count} clients")
            else:
                return self.log_test("Verify Bulk Import", False, f"Expected 4+ clients, found {client_count}")
        else:
            return self.log_test("Verify Bulk Import", False, f"Status: {status}")

    def test_9_create_due_date_federal_service(self):
        """Test 4.1: Create Due Date with Federal Service"""
        if not self.client_id:
            return self.log_test("Create Due Date with Federal Service", False, "No client_id available")
        
        data = {
            "client_id": self.client_id,
            "service_type": "Form 941 - Quarterly Payroll Tax",
            "description": "Q1 2026 Quarterly Payroll Tax Filing",
            "due_date": "2026-04-30T00:00:00Z",
            "is_recurring": True,
            "recurrence_frequency": "quarterly"
        }
        
        success, response, status = self.make_request('POST', 'due-dates', data)
        
        if success and 'id' in response:
            self.due_date_id = response['id']
            return self.log_test("Create Due Date with Federal Service", True)
        else:
            return self.log_test("Create Due Date with Federal Service", False, f"Status: {status}, Response: {response}")

    def test_10_create_due_date_custom_service(self):
        """Test 4.2: Create Due Date with Custom Service"""
        if not self.client_id:
            return self.log_test("Create Due Date with Custom Service", False, "No client_id available")
        
        data = {
            "client_id": self.client_id,
            "service_type": "Business License Renewal - California",
            "description": "Annual business license renewal",
            "due_date": "2026-12-31T00:00:00Z",
            "is_recurring": True,
            "recurrence_frequency": "annually"
        }
        
        success, response, status = self.make_request('POST', 'due-dates', data)
        
        if success and 'id' in response:
            return self.log_test("Create Due Date with Custom Service", True)
        else:
            return self.log_test("Create Due Date with Custom Service", False, f"Status: {status}, Response: {response}")

    def test_11_get_all_due_dates(self):
        """Test 4.3: Get All Due Dates"""
        success, response, status = self.make_request('GET', 'due-dates')
        
        if success and isinstance(response, list):
            # Check that due dates use service_type field (not tax_type)
            has_service_type = all('service_type' in dd for dd in response)
            no_tax_type = all('tax_type' not in dd for dd in response)
            
            if has_service_type and no_tax_type and len(response) >= 2:
                return self.log_test("Get All Due Dates", True, f"Found {len(response)} due dates with service_type")
            else:
                return self.log_test("Get All Due Dates", False, 
                    f"service_type: {has_service_type}, no tax_type: {no_tax_type}, count: {len(response)}")
        else:
            return self.log_test("Get All Due Dates", False, f"Status: {status}")

    def test_12_dashboard_stats(self):
        """Test 5: Dashboard Stats"""
        success, response, status = self.make_request('GET', 'dashboard/stats')
        
        if success and isinstance(response, dict):
            required_fields = ['total_clients', 'total_due_dates', 'upcoming_count', 'overdue_count', 'upcoming_due_dates']
            has_all_fields = all(field in response for field in required_fields)
            
            # Check upcoming_due_dates use service_type
            upcoming = response.get('upcoming_due_dates', [])
            has_service_type = all('service_type' in dd for dd in upcoming)
            
            if has_all_fields and has_service_type:
                return self.log_test("Dashboard Stats", True, 
                    f"Clients: {response['total_clients']}, Due dates: {response['total_due_dates']}")
            else:
                return self.log_test("Dashboard Stats", False, 
                    f"Fields: {has_all_fields}, service_type in upcoming: {has_service_type}")
        else:
            return self.log_test("Dashboard Stats", False, f"Status: {status}")

    def test_13_delete_custom_service_type(self):
        """Test 2.4: Delete Custom Service Type"""
        if not self.custom_service_id:
            return self.log_test("Delete Custom Service Type", False, "No custom_service_id available")
        
        success, response, status = self.make_request('DELETE', f'service-types/{self.custom_service_id}')
        
        if success:
            return self.log_test("Delete Custom Service Type", True)
        else:
            return self.log_test("Delete Custom Service Type", False, f"Status: {status}, Response: {response}")

    def run_all_tests(self):
        """Run all backend API tests"""
        print("🚀 Starting DueDate Backend API Tests")
        print("=" * 50)
        
        # Authentication Tests
        print("\n📋 Authentication & Organization Tests")
        self.test_1_register_with_organization()
        self.test_2_login()
        
        # Service Types Tests
        print("\n🏷️  Service Types Tests (NEW FEATURE)")
        self.test_3_get_service_types()
        self.test_4_create_custom_service_type()
        self.test_5_get_service_types_after_custom()
        
        # Client Management Tests
        print("\n👥 Client Management Tests")
        self.test_6_create_individual_client()
        self.test_7_excel_bulk_upload()
        self.test_8_verify_bulk_import()
        
        # Due Dates Tests
        print("\n📅 Due Dates with Service Types Tests")
        self.test_9_create_due_date_federal_service()
        self.test_10_create_due_date_custom_service()
        self.test_11_get_all_due_dates()
        
        # Dashboard Tests
        print("\n📊 Dashboard Tests")
        self.test_12_dashboard_stats()
        
        # Cleanup Tests
        print("\n🧹 Cleanup Tests")
        self.test_13_delete_custom_service_type()
        
        # Final Results
        print("\n" + "=" * 50)
        print(f"📊 BACKEND TEST RESULTS")
        print(f"Tests Run: {self.tests_run}")
        print(f"Tests Passed: {self.tests_passed}")
        print(f"Tests Failed: {self.tests_run - self.tests_passed}")
        print(f"Success Rate: {(self.tests_passed/self.tests_run)*100:.1f}%")
        
        if self.tests_passed == self.tests_run:
            print("🎉 ALL BACKEND TESTS PASSED!")
            return True
        else:
            print("⚠️  SOME BACKEND TESTS FAILED!")
            return False

def main():
    tester = DueDateAPITester()
    success = tester.run_all_tests()
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())