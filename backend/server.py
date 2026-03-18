from fastapi import FastAPI, APIRouter, HTTPException, Depends, BackgroundTasks, status, UploadFile, File, Header
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from jwt.exceptions import ExpiredSignatureError, InvalidTokenError
from openpyxl import load_workbook
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 168

app = FastAPI()
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# NocoDB Configuration
NOCODB_URL = os.environ.get('NOCODB_URL', 'https://app.nocodb.com')
NOCODB_API_TOKEN = os.environ.get('NOCODB_API_TOKEN', '')
NOCODB_BASE_ID = os.environ.get('NOCODB_BASE_ID', '')

# Table IDs
NOCODB_NOCODB_TABLE_ORGANIZATIONS = os.environ.get('NOCODB_NOCODB_TABLE_ORGANIZATIONS', '')
NOCODB_NOCODB_TABLE_USERS = os.environ.get('NOCODB_NOCODB_TABLE_USERS', '')
NOCODB_NOCODB_TABLE_CLIENTS = os.environ.get('NOCODB_NOCODB_TABLE_CLIENTS', '')
NOCODB_NOCODB_TABLE_DUEDATES = os.environ.get('NOCODB_NOCODB_TABLE_DUEDATES', '')
NOCODB_NOCODB_TABLE_SERVICETYPES = os.environ.get('NOCODB_NOCODB_TABLE_SERVICETYPES', '')
NOCODB_NOCODB_TABLE_REMINDERSETTINGS = os.environ.get('NOCODB_NOCODB_TABLE_REMINDERSETTINGS', '')
NOCODB_NOCODB_TABLE_NOTIFICATIONLOGS = os.environ.get('NOCODB_NOCODB_TABLE_NOTIFICATIONLOGS', '')

# Default service types
DEFAULT_SERVICE_TYPES = {
    'federal': ['Form 941', 'Form 940', 'Form 1120', 'Form 1065', 'Form W-2', 'Form 1099-NEC'],
    'state': ['State Tax Filing', 'State Registration', 'Annual Report'],
    'payroll': ['Payroll Tax', 'Payroll Report', 'W-2 Filing'],
    'custom': [],
    'other': ['Other Compliance']
}

# Pydantic Models
class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    organization_name: str = "My Organization"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ClientCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    company: Optional[str] = None
    notes: Optional[str] = None

class ServiceTypeCreate(BaseModel):
    name: str
    category: str = "custom"

class DueDateCreate(BaseModel):
    client_id: str
    service_type: str
    description: str
    due_date: str
    is_recurring: bool = False
    recurrence_frequency: Optional[str] = None

# Helper Functions
def create_jwt_token(data: dict) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode = data.copy()
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        scheme, token = authorization.split()
        if scheme.lower() != "bearer":
            raise HTTPException(status_code=401, detail="Invalid authentication scheme")
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def get_nocodb_headers():
    return {
        "xc-token": NOCODB_API_TOKEN,
        "Content-Type": "application/json"
    }

def get_nocodb_url(endpoint: str) -> str:
    return f"{NOCODB_URL}{endpoint}"

# NocoDB API Helper
import httpx

async def nocodb_get(endpoint: str, params: dict = None):
    url = get_nocodb_url(endpoint)
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(url, headers=get_nocodb_headers(), params=params, timeout=30.0)
            if response.status_code >= 400:
                logging.error(f"NocoDB GET error: {response.status_code} - {response.text}")
                return None
            return response.json()
        except Exception as e:
            logging.error(f"NocoDB GET exception: {e}")
            return None

async def nocodb_post(endpoint: str, json_data: dict):
    url = get_nocodb_url(endpoint)
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(url, headers=get_nocodb_headers(), json=json_data, timeout=30.0)
            if response.status_code >= 400:
                logging.error(f"NocoDB POST error: {response.status_code} - {response.text}")
                return None
            return response.json()
        except Exception as e:
            logging.error(f"NocoDB POST exception: {e}")
            return None

async def nocodb_patch(endpoint: str, json_data: dict):
    url = get_nocodb_url(endpoint)
    async with httpx.AsyncClient() as client:
        try:
            response = await client.patch(url, headers=get_nocodb_headers(), json=json_data, timeout=30.0)
            if response.status_code >= 400:
                logging.error(f"NocoDB PATCH error: {response.status_code} - {response.text}")
                return None
            return response.json()
        except Exception as e:
            logging.error(f"NocoDB PATCH exception: {e}")
            return None

async def nocodb_delete(endpoint: str, json_data: dict):
    url = get_nocodb_url(endpoint)
    async with httpx.AsyncClient() as client:
        try:
            response = await client.delete(url, headers=get_nocodb_headers(), json=json_data, timeout=30.0)
            if response.status_code >= 400:
                logging.error(f"NocoDB DELETE error: {response.status_code} - {response.text}")
                return None
            return response.json()
        except Exception as e:
            logging.error(f"NocoDB DELETE exception: {e}")
            return None

# Auth Endpoints
@api_router.post("/auth/register")
async def register(data: UserRegister):
    try:
        # Create organization
        org_id = str(uuid.uuid4())
        org_data = {
            "id": org_id,
            "name": data.organization_name,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        if NOCODB_TABLE_ORGANIZATIONS:
            await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_ORGANIZATIONS}/records", org_data)
        else:
            # Store in memory if NocoDB not configured
            pass
        
        # Create user
        user_id = str(uuid.uuid4())
        hashed_password = pwd_context.hash(data.password)
        user_data = {
            "id": user_id,
            "email": data.email,
            "name": data.name,
            "password": hashed_password,
            "organization_id": org_id,
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        if NOCODB_TABLE_USERS:
            await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_USERS}/records", user_data)
        
        # Generate token
        token = create_jwt_token({
            "user_id": user_id,
            "email": data.email,
            "organization_id": org_id
        })
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user_id,
                "email": data.email,
                "name": data.name,
                "role": "admin"
            },
            "organization": {
                "id": org_id,
                "name": data.organization_name
            }
        }
    except Exception as e:
        logging.error(f"Registration error: {e}")
        raise HTTPException(status_code=500, detail="Registration failed")

@api_router.post("/auth/login")
async def login(data: UserLogin):
    try:
        # Find user in NocoDB
        if NOCODB_TABLE_USERS:
            result = await nocodb_get(
                f"/api/v2/tables/{NOCODB_TABLE_USERS}/records",
                params={"where": f"(email,eq,{data.email})", "limit": 1}
            )
            
            if result and result.get('list'):
                user = result['list'][0]
            else:
                raise HTTPException(status_code=401, detail="Invalid credentials")
        else:
            raise HTTPException(status_code=500, detail="Database not configured")
        
        # Verify password
        if not pwd_context.verify(data.password, user.get('password', '')):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Generate token
        token = create_jwt_token({
            "user_id": user['id'],
            "email": user['email'],
            "organization_id": user['organization_id']
        })
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user['id'],
                "email": user['email'],
                "name": user['name'],
                "role": user.get('role', 'member')
            },
            "organization": {
                "id": user['organization_id'],
                "name": user.get('organization_name', 'My Organization')
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")

# Dashboard
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    stats = {
        "total_clients": 0,
        "total_due_dates": 0,
        "upcoming_count": 0,
        "overdue_count": 0,
        "upcoming_due_dates": []
    }
    
    now = datetime.now(timezone.utc)
    thirty_days = now + timedelta(days=30)
    
    # Get clients count
    if NOCODB_TABLE_CLIENTS and org_id:
        result = await nocodb_get(
            f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records",
            params={"where": f"(organization_id,eq,{org_id})", "limit": 1000}
        )
        if result:
            stats["total_clients"] = len(result.get('list', []))
    
    # Get due dates
    if NOCODB_TABLE_DUEDATES and org_id:
        result = await nocodb_get(
            f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records",
            params={"where": f"(organization_id,eq,{org_id})", "limit": 1000}
        )
        if result:
            due_dates = result.get('list', [])
            stats["total_due_dates"] = len(due_dates)
            
            for dd in due_dates:
                due_date = datetime.fromisoformat(dd['due_date'].replace('Z', '+00:00')) if isinstance(dd['due_date'], str) else dd['due_date']
                
                if due_date < now:
                    stats["overdue_count"] += 1
                elif due_date <= thirty_days:
                    stats["upcoming_count"] += 1
                    # Enrich with client info
                    dd['client'] = None
                    if NOCODB_TABLE_CLIENTS:
                        client_result = await nocodb_get(
                            f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records",
                            params={"where": f"(id,eq,{dd.get('client_id')})", "limit": 1}
                        )
                        if client_result and client_result.get('list'):
                            dd['client'] = client_result['list'][0]
                    stats["upcoming_due_dates"].append(dd)
    
    return stats

# Clients
@api_router.get("/clients")
async def get_clients(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    if NOCODB_TABLE_CLIENTS and org_id:
        result = await nocodb_get(
            f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records",
            params={"where": f"(organization_id,eq,{org_id})", "limit": 1000}
        )
        return result.get('list', []) if result else []
    return []

@api_router.post("/clients")
async def create_client(data: ClientCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    client_id = str(uuid.uuid4())
    client_data = {
        "id": client_id,
        "organization_id": org_id,
        "name": data.name,
        "email": data.email,
        "phone": data.phone,
        "company": data.company,
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    if NOCODB_TABLE_CLIENTS:
        result = await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records", client_data)
        if result:
            return {"id": client_id, **data.model_dump()}
    
    return {"id": client_id, **data.model_dump()}

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, data: ClientCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    update_data = {
        "Id": client_id,
        "name": data.name,
        "email": data.email,
        "phone": data.phone,
        "company": data.company,
        "notes": data.notes
    }
    
    if NOCODB_TABLE_CLIENTS:
        await nocodb_patch(f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records", update_data)
    
    return {"id": client_id, **data.model_dump()}

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    
    if NOCODB_TABLE_CLIENTS:
        await nocodb_delete(f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records", {"Id": client_id})
    
    return {"message": "Client deleted"}

@api_router.post("/clients/upload-excel")
async def upload_clients_excel(file: UploadFile = File(...), authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    
    contents = await file.read()
    
    try:
        workbook = load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        # Skip header row
        headers = [cell.value for cell in sheet[1]]
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[headers.index('Name')] if 'Name' in headers else row[0]
                email = row[headers.index('Email')] if 'Email' in headers else row[1]
                phone = row[headers.index('Phone')] if 'Phone' in headers else None
                company = row[headers.index('Company')] if 'Company' in headers else None
                notes = row[headers.index('Notes')] if 'Notes' in headers else None
                
                if not name or not email:
                    errors.append(f"Row {row_idx}: Missing name or email")
                    continue
                
                client_id = str(uuid.uuid4())
                client_data = {
                    "id": client_id,
                    "organization_id": org_id,
                    "name": str(name),
                    "email": str(email),
                    "phone": str(phone) if phone else None,
                    "company": str(company) if company else None,
                    "notes": str(notes) if notes else None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                if NOCODB_TABLE_CLIENTS:
                    await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_CLIENTS}/records", client_data)
                
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported} clients",
            "imported": imported,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

# Service Types
@api_router.get("/service-types")
async def get_service_types(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    service_types = {k: v.copy() for k, v in DEFAULT_SERVICE_TYPES.items()}
    
    # Get custom service types from NocoDB
    if NOCODB_TABLE_SERVICETYPES and org_id:
        result = await nocodb_get(
            f"/api/v2/tables/{NOCODB_TABLE_SERVICETYPES}/records",
            params={"where": f"(organization_id,eq,{org_id})", "limit": 1000}
        )
        if result:
            for st in result.get('list', []):
                category = st.get('category', 'custom')
                name = st.get('name')
                if category in service_types and name:
                    if name not in service_types[category]:
                        service_types[category].append(name)
    
    return service_types

@api_router.post("/service-types")
async def create_service_type(data: ServiceTypeCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    st_id = str(uuid.uuid4())
    st_data = {
        "id": st_id,
        "organization_id": org_id,
        "name": data.name,
        "category": data.category,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    if NOCODB_TABLE_SERVICETYPES:
        await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_SERVICETYPES}/records", st_data)
    
    return {"id": st_id, **data.model_dump()}

# Due Dates
@api_router.get("/due-dates")
async def get_due_dates(authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    if NOCODB_TABLE_DUEDATES and org_id:
        result = await nocodb_get(
            f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records",
            params={"where": f"(organization_id,eq,{org_id})", "limit": 1000}
        )
        return result.get('list', []) if result else []
    return []

@api_router.post("/due-dates")
async def create_due_date(data: DueDateCreate, authorization: str = Header(None)):
    user = await get_current_user(authorization)
    org_id = user.get("organization_id")
    
    dd_id = str(uuid.uuid4())
    dd_data = {
        "id": dd_id,
        "organization_id": org_id,
        "client_id": data.client_id,
        "service_type": data.service_type,
        "description": data.description,
        "due_date": data.due_date,
        "is_recurring": data.is_recurring,
        "recurrence_frequency": data.recurrence_frequency,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    if NOCODB_TABLE_DUEDATES:
        await nocodb_post(f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records", dd_data)
    
    return {"id": dd_id, **data.model_dump()}

@api_router.put("/due-dates/{dd_id}")
async def update_due_date(dd_id: str, data: DueDateCreate, authorization: str = Header(None)):
    update_data = {
        "Id": dd_id,
        "client_id": data.client_id,
        "service_type": data.service_type,
        "description": data.description,
        "due_date": data.due_date,
        "is_recurring": data.is_recurring,
        "recurrence_frequency": data.recurrence_frequency
    }
    
    if NOCODB_TABLE_DUEDATES:
        await nocodb_patch(f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records", update_data)
    
    return {"id": dd_id, **data.model_dump()}

@api_router.delete("/due-dates/{dd_id}")
async def delete_due_date(dd_id: str, authorization: str = Header(None)):
    if NOCODB_TABLE_DUEDATES:
        await nocodb_delete(f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records", {"Id": dd_id})
    
    return {"message": "Due date deleted"}

@api_router.patch("/due-dates/{dd_id}/status")
async def update_due_date_status(dd_id: str, status: str, authorization: str = Header(None)):
    update_data = {
        "Id": dd_id,
        "status": status
    }
    
    if NOCODB_TABLE_DUEDATES:
        await nocodb_patch(f"/api/v2/tables/{NOCODB_TABLE_DUEDATES}/records", update_data)
    
    return {"message": "Status updated"}

# Health check
@api_router.get("/health")
async def health_check():
    return {"status": "ok"}

# Include router
app.include_router(api_router)

@app.get("/")
async def root():
    return {"message": "DueDate API", "version": "1.0.0"}
